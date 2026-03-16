import openpyxl
import re

def norm_dia(d):
    """Normaliza abreviaciones de días al nombre completo."""
    s = str(d).upper().strip()
    s = s.replace('Á','A').replace('É','E').replace('Í','I').replace('Ó','O').replace('Ú','U')
    mapping = {
        'LU': 'LUNES',   'LUN': 'LUNES',
        'MA': 'MARTES',  'MAR': 'MARTES',
        'MI': 'MIERCOLES', 'MIE': 'MIERCOLES', 'MIÉ': 'MIERCOLES',
        'JU': 'JUEVES',  'JUE': 'JUEVES',
        'VI': 'VIERNES', 'VIE': 'VIERNES',
        'SA': 'SABADO',  'SAB': 'SABADO',  'SÁ': 'SABADO',
        'DO': 'DOMINGO', 'DOM': 'DOMINGO',
    }
    return mapping.get(s, mapping.get(s[:3], s))

def parse_hora(h):
    """Convierte hora a entero HHMM. Soporta float, int y string."""
    if h is None or h == '':
        return None
    if isinstance(h, (int, float)):
        v = float(h)
        if 0 < v < 1:
            total = round(v * 24 * 60)
            return (total // 60) * 100 + (total % 60)
        if 0 <= v <= 24:
            return int(v) * 100
    s = str(h).replace(':', '').replace('.', '').strip()
    if not s or not s.isdigit():
        return None
    s = s[:4]
    if len(s) <= 2:
        return int(s) * 100
    if len(s) == 3:
        return int(s[0]) * 100 + int(s[1:])
    return int(s[:2]) * 100 + int(s[2:4])

def norm_str(v):
    """Limpia un string: quita nbsp, espacios extra."""
    if v is None:
        return ''
    # Reemplaza espacios de no ruptura por espacios normales y limpia extremos
    return ' '.join(str(v).replace('\u00a0', ' ').replace('\xa0', ' ').split()).strip()

def find_header_row(ws):
    """Encuentra la fila que contiene los encabezados reales."""
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        row_vals = [norm_str(c).upper() for c in row if c]
        # Busca palabras clave para identificar la cabecera
        if any(keyword in row_vals for keyword in ['CÓDIGO', 'CODIGO', 'NOMBRE DEL CURSO', 'ASIGNATURA']):
            return i
    return None

def get_col_indices(header_row):
    """Mapea nombre de columna a índice, tolerante a variaciones."""
    idx = {}
    for i, val in enumerate(header_row):
        if val is None:
            continue
        key = norm_str(val).upper()
        
        if re.search(r'C[ÓO]DIGO', key):
            idx['codigo'] = i
        elif 'NOMBRE DEL CURSO' in key or 'ASIGNATURA' in key:
            idx['nombre'] = i
        elif re.search(r'SECCI[ÓO]N', key) or 'GRUPO' in key:
            idx['seccion'] = i
        elif (re.search(r'APELLIDOS', key) and 'DOCENTE' in key) or 'PROFESOR' in key:
            idx['docente'] = i
        elif key.strip().startswith('TIPO'):
            idx['tipo'] = i
        elif 'AULA' in key:
            idx['aula'] = i
        elif re.search(r'D[ÍI]A', key):
            idx['dia'] = i
        elif 'INICIO' in key or 'H. INICIO' in key:
            idx['inicio'] = i
        elif re.search(r'FINAL|FIN\b|H. FIN', key):
            idx['fin'] = i
    return idx

def parsear_excel(file_obj):
    """Parsea la carga horaria desde un objeto de archivo Excel."""
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    header_row_num = find_header_row(ws)
    if header_row_num is None:
        raise ValueError("No se encontró la fila de encabezados. Asegúrate de que el Excel tenga columnas como 'Curso' o 'Código'.")

    rows = list(ws.iter_rows(values_only=True))
    header = rows[header_row_num - 1]
    col = get_col_indices(header)

    # Columnas críticas necesarias para que el generador funcione
    required = ['nombre', 'seccion', 'dia', 'inicio', 'fin']
    missing = [r for r in required if r not in col]
    if missing:
        raise ValueError(f"Faltan columnas obligatorias: {missing}")

    carga = {}

    for row in rows[header_row_num:]:
        if not any(row): continue

        nombre  = norm_str(row[col['nombre']]) if 'nombre' in col else ''
        seccion = norm_str(row[col['seccion']]) if 'seccion' in col else ''
        docente = norm_str(row[col['docente']]) if 'docente' in col else 'POR ASIGNAR'
        tipo    = norm_str(row[col['tipo']]).upper() if 'tipo' in col else 'P'
        dia_raw = row[col['dia']] if 'dia' in col else ''
        inicio  = row[col['inicio']]
        fin_val = row[col['fin']]
        aula    = norm_str(row[col['aula']]) if 'aula' in col else 'S/A'

        if not nombre or not seccion or not dia_raw: continue

        dia = norm_dia(dia_raw)
        ini = parse_hora(inicio)
        fin = parse_hora(fin_val)

        if ini is None or fin is None or ini >= fin: continue

        if nombre not in carga:
            carga[nombre] = {}
        if seccion not in carga[nombre]:
            carga[nombre][seccion] = {'docente': docente, 'clases': []}

        carga[nombre][seccion]['clases'].append({
            'dia': dia, 'ini': ini, 'fin': fin,
            'tipo': tipo, 'aula': aula
        })

    return carga